# On utilise Pandas pour manipuler les exports des différents logiciels (Orbis, Hexagone)

import pandas as pd
import numpy as np
import logging
import argparse
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ==========================================
# CONFIGURATION GLOBALE
# ==========================================
DOSSIER_IMPORT = Path('./data_test/import_test')
DOSSIER_EXPORT = Path('./data_test/export_test')
DOSSIER_EXPORT.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    filename=DOSSIER_EXPORT / f'Logs_Conciliation_{datetime.now().strftime("%Y%m%d")}.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Colonnes attendues par type de fichier (pour validation)
COLONNES_ORBIS = ['N° Hospit', 'Entrée le', 'Sortie le', 'Nom', 'Prénom', 'Né(e) le', 'UM', 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM']
COLONNES_HEXA_HOSPIT = ['Date', 'Nom/Prénom', 'Date de naissance', 'Dossier', 'Date entrée']
COLONNES_HEXA_SEANCES = ['Nom/Prénom', 'Date de naissance', 'N° Dossier', 'Date']

# ==========================================
# FONCTIONS UTILITAIRES
# ==========================================
def valider_colonnes(df, colonnes_attendues, nom_fichier):
    """
    Sanity check : vérifie que le DataFrame contient bien les colonnes dont on a besoin.
    Si le format de l'export source change (ce qui arrive souvent), ça permet de planter proprement
    au lieu d'avoir des erreurs illisibles plus bas dans le code.
    """
    # On passe par des sets (ensembles) pour faire la comparaison facilement
    colonnes_presentes = set(df.columns)
    colonnes_manquantes = [c for c in colonnes_attendues if c not in colonnes_presentes]
    
    if colonnes_manquantes:
        # On arrête tout en levant une exception explicite
        msg = (f"Le fichier '{nom_fichier}' ne contient pas les colonnes attendues.\n"
               f"  Colonnes manquantes : {colonnes_manquantes}\n"
               f"  Colonnes trouvees   : {list(df.columns)}")
        logging.error(msg)
        raise ValueError(msg)
        
    logging.info(f"Validation OK pour '{nom_fichier}' ({len(df)} lignes, {len(df.columns)} colonnes)")

def nettoyer_nda_hexa(nda_series):
    """
    Nettoie la colonne des numéros de dossier.
    Hexagone rajoute parfois des lettres avant le NDA (genre 'H 12345').
    On utilise une regex simple pour dégager une majuscule + un espace en début de string.
    """
    return nda_series.astype(str).str.replace(r'^[A-Z]\s', '', regex=True).str.strip()

def formater_date_jjmmaaaa(date_series):
    """
    Normalise les dates en format string JJ/MM/AAAA.
    Les imports Excel peuvent ramener des types mixtes (datetime ou texte). Pour nos jointures, 
    on a besoin d'un format texte 100% fiable.
    """
    # 'coerce' gère les trucs impossibles à parser (ça renvoie NaT au lieu de planter)
    dt_obj = pd.to_datetime(date_series, dayfirst=True, errors='coerce')
    
    # On formatte en texte et on met au propre les valeurs nulles pour éviter les conflits plus tard
    res = dt_obj.dt.strftime('%d/%m/%Y')
    return res.replace(['NaT', 'nan', 'NaN'], np.nan)

def charger_famille_hexagone(dossier, mots_cles, colonnes_validation):
    """Charge et fusionne plusieurs fichiers Hexagone selon une liste de mots-clés."""
    dfs = []
    fichiers_deja_charges = set()
    for mot in mots_cles:
        fichiers = [f for f in dossier.iterdir() if f.is_file() and mot.lower() in f.name.lower() and not f.name.startswith('~')]
        for f in fichiers:
            chemin_normalise = f.resolve()
            if chemin_normalise in fichiers_deja_charges:
                continue
            try:
                # Les fichiers Hexagone ont 2 lignes d'en-tête (Titre du rapport, ligne vide) avant les vraies colonnes
                df = pd.read_excel(f, dtype=str, header=2)
                valider_colonnes(df, colonnes_validation, f.name)
                df['Fichier Source'] = f.name
                logging.info(f"Fichier Hexagone chargé : {f.name}")
                dfs.append(df)
                fichiers_deja_charges.add(chemin_normalise)
            except Exception as e:
                logging.error(f"Erreur lecture {f.name} : {e}")
                raise
    
    if not dfs:
        raise FileNotFoundError(f"Aucun fichier Hexagone trouvé pour les mots-clés : {mots_cles}")
    
    return pd.concat(dfs, ignore_index=True)

def mettre_en_forme_excel(chemin_fichier):
    """Applique la mise en forme sur un fichier Excel existant :
    - En-têtes en gras avec fond coloré
    - Largeur de colonnes auto-ajustée
    - Filtres automatiques activés"""
    from openpyxl import load_workbook
    
    wb = load_workbook(chemin_fichier)
    
    # Style des en-têtes
    style_entete = Font(bold=True, color="FFFFFF", size=11)
    fond_entete = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    alignement = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue
            
        # Mise en forme des en-têtes (ligne 1)
        for cell in ws[1]:
            cell.font = style_entete
            cell.fill = fond_entete
            cell.alignment = alignement
        
        # Auto-ajustement de la largeur des colonnes
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            # Largeur = max + marge, plafonnée à 40
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)
        
        # Activation des filtres automatiques
        ws.auto_filter.ref = ws.dimensions
    
    wb.save(chemin_fichier)

# ==========================================
# FONCTION PRINCIPALE
# ==========================================
def lancer_conciliation(orbis_path=None, hexa_hospit_paths=None, hexa_seances_paths=None, export_dir=None):
    """Lance la conciliation. Si les chemins sont fournis (mode GUI), les utilise directement.
    Sinon, recherche automatiquement les fichiers dans le dossier courant (mode CLI)."""
    
    if export_dir:
        dossier_export = Path(export_dir)
        dossier_export.mkdir(parents=True, exist_ok=True)
    else:
        dossier_export = DOSSIER_EXPORT
    
    logging.info("--- DÉMARRAGE DU TRAITEMENT ---")

    # ---------------------------------------------------------
    # 1. PRÉPARATION DU FICHIER ORBIS
    # ---------------------------------------------------------
    if orbis_path:
        chemin_orbis = Path(orbis_path)
    else:
        fichiers_orbis = [f for f in DOSSIER_IMPORT.iterdir() if f.is_file() and 'orbis' in f.name.lower() and not f.name.startswith('~')]
        if not fichiers_orbis:
            raise FileNotFoundError("Fichier Orbis introuvable.")
        chemin_orbis = fichiers_orbis[0]
    
    # On force la lecture en texte (dtype=str) pour éviter que Pandas fasse de l'auto-détection foireuse
    # et supprime par exemple les 0 au début des numéros de dossier.
    df_orbis_brut = pd.read_excel(chemin_orbis, dtype=str)
    valider_colonnes(df_orbis_brut, COLONNES_ORBIS, chemin_orbis.name)
    
    # --- Règle métier --- 
    # On filtre les données : on garde toutes les lignes où l'UM ne contient PAS '1402' 
    # (le tilde '~' est l'opérateur logique NOT en Pandas).
    df_orbis = df_orbis_brut[~df_orbis_brut['UM'].astype(str).str.contains('1402', na=False)].copy()

    # Nettoyage des identifiants et des dates pour préparer la jointure
    df_orbis['N° Hospit'] = df_orbis['N° Hospit'].str.strip()
    df_orbis['Entrée le'] = formater_date_jjmmaaaa(df_orbis['Entrée le'])
    df_orbis['Sortie le'] = formater_date_jjmmaaaa(df_orbis['Sortie le'])
    df_orbis['Entrée le.1'] = formater_date_jjmmaaaa(df_orbis['Entrée le.1']) # Orbis nomme la date de venue ainsi
    
    # --- Séparation des flux selon les Unités Médicales (UM) ---
    # Les règles de rapprochement sont différentes entre les séances et l'hospit, donc on sépare.
    ums_seances = ['1400', '1401', '1048']
    
    # .isin() permet de filtrer facilement sur une liste de valeurs (ce qui permet de guarder uniquement les UM séances)
    df_orbis_seances_filtre = df_orbis[df_orbis['UM'].astype(str).str.strip().isin(ums_seances)].copy()
    
    # On exclut les UM de séances
    df_orbis_hospit_filtre = df_orbis[~df_orbis['UM'].astype(str).str.strip().isin(ums_seances)].copy()
    
    # Pour l'hospitalisation, on ne traite qu'une ligne par séjour.
    # On dé-doublonne sur le NDA en gardant la première occurrence.
    df_orbis_hospit_sejour = df_orbis_hospit_filtre.drop_duplicates(subset=['N° Hospit'], keep='first')
    # ---------------------------------------------------------

    # ---------------------------------------------------------
    # 2. PRÉPARATION DES FICHIERS HEXAGONE (HOSPITALISATIONS)
    # ---------------------------------------------------------
    if hexa_hospit_paths:
        # Mode GUI : charger les fichiers spécifiés directement
        dfs_hospit = []
        for p in hexa_hospit_paths:
            fp = Path(p)
            df = pd.read_excel(fp, dtype=str, header=2)
            valider_colonnes(df, COLONNES_HEXA_HOSPIT, fp.name)
            df['Fichier Source'] = fp.name
            dfs_hospit.append(df)
        df_hexa_hospit = pd.concat(dfs_hospit, ignore_index=True)
    elif hexa_hospit_paths is not None:
        # Mode GUI : aucun fichier sélectionné
        df_hexa_hospit = pd.DataFrame()
    else:
        mots_hospit = ['hospit', 'nouveaune', 'nouveau', 'ortho']
        df_hexa_hospit = charger_famille_hexagone(DOSSIER_IMPORT, mots_hospit, COLONNES_HEXA_HOSPIT)
    
    if not df_hexa_hospit.empty:
        df_hexa_hospit['NDA'] = nettoyer_nda_hexa(df_hexa_hospit['Dossier'])
        df_hexa_hospit[['Nom', 'Prénom']] = df_hexa_hospit['Nom/Prénom'].str.split('/', n=1, expand=True)
        df_hexa_hospit['Date entrée'] = formater_date_jjmmaaaa(df_hexa_hospit['Date entrée'])
        df_hexa_hospit['Date de sortie'] = formater_date_jjmmaaaa(df_hexa_hospit['Date'])

    # ---------------------------------------------------------
    # 3. PRÉPARATION DES FICHIERS HEXAGONE (SÉANCES)
    # ---------------------------------------------------------
    if hexa_seances_paths:
        # Mode GUI : charger les fichiers spécifiés directement
        dfs_seances = []
        for p in hexa_seances_paths:
            fp = Path(p)
            df = pd.read_excel(fp, dtype=str, header=2)
            valider_colonnes(df, COLONNES_HEXA_SEANCES, fp.name)
            df['Fichier Source'] = fp.name
            dfs_seances.append(df)
        df_hexa_seances = pd.concat(dfs_seances, ignore_index=True)
    elif hexa_seances_paths is not None:
        # Mode GUI : aucun fichier sélectionné
        df_hexa_seances = pd.DataFrame()
    else:
        mots_seances = ['chimio', 'dialyse']
        df_hexa_seances = charger_famille_hexagone(DOSSIER_IMPORT, mots_seances, COLONNES_HEXA_SEANCES)
    
    if not df_hexa_seances.empty:
        df_hexa_seances['NDA'] = df_hexa_seances['N° Dossier'].astype(str).str.strip()
        df_hexa_seances[['Nom', 'Prénom']] = df_hexa_seances['Nom/Prénom'].str.split('/', n=1, expand=True)
        df_hexa_seances['Date de venue'] = formater_date_jjmmaaaa(df_hexa_seances['Date'])

    # =========================================================
    # TRI 1 : COMPARAISON FAMILLE HOSPITALISATION
    # =========================================================
    logging.info("Exécution Tri 1 : Différences Hospitalisations (Clé: NDA)")
    
    cols_orbis_t1 = ['N° Hospit', 'Nom', 'Prénom', 'Né(e) le', 'Entrée le', 'Sortie le', 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM']
    cols_hexa_t1 = ['NDA', 'Nom', 'Prénom', 'Date de naissance', 'Date entrée', 'Date de sortie']
    
    if not df_hexa_hospit.empty:
        # Utilisation de df_orbis_hospit_sejour au lieu du fichier complet
        # --- Jointure Externe (Outer Join) ---
        # L'idée est de croiser les deux tables sur le numéro de dossier.
        # how='outer' signifie qu'on garde toutes les lignes (celles qui matchent ET celles qui sont orphelines).
        # indicator=True ajoute une colonne '_merge' pour nous dire d'où vient chaque ligne ('both', 'left_only', 'right_only').
        tri1 = pd.merge(
            df_orbis_hospit_sejour[cols_orbis_t1], 
            df_hexa_hospit[cols_hexa_t1], 
            left_on='N° Hospit', right_on='NDA', 
            how='outer', indicator=True
        )
        
        # Les anomalies sont tout simplement les lignes qui ne sont pas dans 'both' (donc soit dans l'un, soit dans l'autre)
        anomalies_tri1 = tri1[tri1['_merge'] != 'both'].copy()
    else:
        anomalies_tri1 = pd.DataFrame()
    
    # Colonne indiquant l'origine de l'écart
    if not anomalies_tri1.empty:
        anomalies_tri1['Origine de l\'écart'] = anomalies_tri1['_merge'].map({
            'left_only': 'Manquant dans Hexagone',
            'right_only': 'Manquant dans Orbis'
        })
    
    if not anomalies_tri1.empty:
        anomalies_tri1['NDA Final'] = anomalies_tri1['N° Hospit'].combine_first(anomalies_tri1['NDA'])
        anomalies_tri1['Nom Final'] = anomalies_tri1['Nom_x'].combine_first(anomalies_tri1['Nom_y'])
        anomalies_tri1['Prénom Final'] = anomalies_tri1['Prénom_x'].combine_first(anomalies_tri1['Prénom_y'])
        anomalies_tri1['Date Naissance Final'] = anomalies_tri1['Né(e) le'].combine_first(anomalies_tri1['Date de naissance'])
        anomalies_tri1['Date Entrée Final'] = anomalies_tri1['Entrée le'].combine_first(anomalies_tri1['Date entrée'])
        anomalies_tri1['Date Sortie Final'] = anomalies_tri1['Sortie le'].combine_first(anomalies_tri1['Date de sortie'])
    
    colonnes_export_tri1 = [
        'NDA Final', 'Nom Final', 'Prénom Final', 'Date Naissance Final', 
        'Date Entrée Final', 'Date Sortie Final', 'Origine de l\'écart', 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM'
    ]
    if not anomalies_tri1.empty:
        export_tri1 = anomalies_tri1[colonnes_export_tri1].rename(columns={'NDA Final': 'NDA', 'Nom Final': 'Nom', 'Prénom Final': 'Prénom', 'Date Naissance Final': 'Date Naissance', 'Date Entrée Final': 'Date entrée', 'Date Sortie Final': 'Date sortie'})
        
        # Tri par date d'entrée décroissante
        export_tri1['_tri_date'] = pd.to_datetime(export_tri1['Date entrée'], dayfirst=True, errors='coerce')
        export_tri1 = export_tri1.sort_values('_tri_date', ascending=False).drop(columns=['_tri_date'])
    else:
        export_tri1 = pd.DataFrame(columns=['NDA', 'Nom', 'Prénom', 'Date Naissance', 'Date entrée', 'Date sortie', "Origine de l'écart", 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM'])

    # =========================================================
    # TRI 2 : COMPARAISON FAMILLE SÉANCES
    # =========================================================
    logging.info("Exécution Tri 2 : Différences Séances (Clé: NDA + Date de venue)")
    
    cols_orbis_t2 = ['N° Hospit', 'Nom', 'Prénom', 'Né(e) le', 'Entrée le.1', 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM']
    cols_hexa_t2 = ['NDA', 'Nom', 'Prénom', 'Date de naissance', 'Date de venue']
    
    if not df_hexa_seances.empty:
        # Utilisation de df_orbis_seances_filtre au lieu du fichier complet
        # Pareil que pour le Tri 1, mais ici la clé de jointure est composite : 
        # il faut que le NDA ET la date de venue correspondent pour valider une séance.
        tri2 = pd.merge(
            df_orbis_seances_filtre[cols_orbis_t2], 
            df_hexa_seances[cols_hexa_t2], 
            left_on=['N° Hospit', 'Entrée le.1'], right_on=['NDA', 'Date de venue'], 
            how='outer', indicator=True
        )
        
        anomalies_tri2 = tri2[tri2['_merge'] != 'both'].copy()
    else:
        anomalies_tri2 = pd.DataFrame()
    
    # Colonne indiquant l'origine de l'écart
    if not anomalies_tri2.empty:
        anomalies_tri2['Origine de l\'écart'] = anomalies_tri2['_merge'].map({
            'left_only': 'Manquant dans Hexagone',
            'right_only': 'Manquant dans Orbis'
        })
    
    if not anomalies_tri2.empty:
        anomalies_tri2['NDA Final'] = anomalies_tri2['N° Hospit'].combine_first(anomalies_tri2['NDA'])
        anomalies_tri2['Nom Final'] = anomalies_tri2['Nom_x'].combine_first(anomalies_tri2['Nom_y'])
        anomalies_tri2['Prénom Final'] = anomalies_tri2['Prénom_x'].combine_first(anomalies_tri2['Prénom_y'])
        anomalies_tri2['Date Naissance Final'] = anomalies_tri2['Né(e) le'].combine_first(anomalies_tri2['Date de naissance'])
        anomalies_tri2['Date Venue Final'] = anomalies_tri2['Entrée le.1'].combine_first(anomalies_tri2['Date de venue'])

    colonnes_export_tri2 = [
        'NDA Final', 'Nom Final', 'Prénom Final', 'Date Naissance Final', 
        'Date Venue Final', 'Origine de l\'écart', 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM'
    ]
    if not anomalies_tri2.empty:
        export_tri2 = anomalies_tri2[colonnes_export_tri2].rename(columns={'NDA Final': 'NDA', 'Nom Final': 'Nom', 'Prénom Final': 'Prénom', 'Date Naissance Final': 'Date Naissance', 'Date Venue Final': 'Date de venue'})
        
        # Tri par date de venue décroissante
        export_tri2['_tri_date'] = pd.to_datetime(export_tri2['Date de venue'], dayfirst=True, errors='coerce')
        export_tri2 = export_tri2.sort_values('_tri_date', ascending=False).drop(columns=['_tri_date'])
    else:
        export_tri2 = pd.DataFrame(columns=['NDA', 'Nom', 'Prénom', 'Date Naissance', 'Date de venue', "Origine de l'écart", 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM'])

    # =========================================================
    # TRI 3 : SANS DATE DE SORTIE (HOSPITALISATIONS)
    # =========================================================
    logging.info("Exécution Tri 3 : Analyse des dates de sorties manquantes")
    
    if not df_hexa_hospit.empty:
        # Utilisation de df_orbis_hospit_sejour ici aussi
        # --- Jointure Interne (Inner Join) ---
        # Ici, on ne cherche pas les orphelins, on veut juste regarder les patients qui sont bien dans les 2 systèmes
        # pour vérifier s'il nous manque une date de sortie.
        tri3 = pd.merge(
            df_orbis_hospit_sejour[['N° Hospit', 'Nom', 'Prénom', 'Né(e) le', 'Entrée le', 'Sortie le', 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM']], 
            df_hexa_hospit[['NDA', 'Date de sortie', 'Fichier Source']], 
            left_on='N° Hospit', right_on='NDA', 
            how='inner' 
        )
        
        # On filtre les lignes où l'une des deux dates de sortie est nulle (isna).
        # L'opérateur '|' est le OU (OR) vectorisé de Pandas.
        anomalies_tri3 = tri3[pd.isna(tri3['Sortie le']) | pd.isna(tri3['Date de sortie'])].copy()
    else:
        anomalies_tri3 = pd.DataFrame()
    
    # "Date trouvée ailleurs ?" : Oui si au moins un logiciel a la date, Non sinon
    if not anomalies_tri3.empty:
        cond_manque_orbis = pd.isna(anomalies_tri3['Sortie le']) & pd.notna(anomalies_tri3['Date de sortie'])
        cond_manque_hexa = pd.notna(anomalies_tri3['Sortie le']) & pd.isna(anomalies_tri3['Date de sortie'])
        cond_manque_partout = pd.isna(anomalies_tri3['Sortie le']) & pd.isna(anomalies_tri3['Date de sortie'])
        anomalies_tri3['Date trouvée ailleurs ?'] = np.select(
            [cond_manque_orbis, cond_manque_hexa, cond_manque_partout],
            ['Oui', 'Oui', 'Non'],
            default='Inconnu'
        )
        
        # "Source de la date" : nom exact du fichier Hexa, "Orbis", ou "Aucun"
        anomalies_tri3['Source de la date'] = np.select(
            [cond_manque_orbis, cond_manque_hexa, cond_manque_partout],
            [anomalies_tri3['Fichier Source'].values, 'Orbis', 'Aucun'],
            default='Inconnu'
        )
        
        anomalies_tri3['Date Sortie Consolidée'] = anomalies_tri3['Sortie le'].combine_first(anomalies_tri3['Date de sortie'])

    colonnes_export_tri3 = [
        'N° Hospit', 'Nom', 'Prénom', 'Né(e) le', 'Entrée le', 'Date Sortie Consolidée', 
        'Date trouvée ailleurs ?', 'Source de la date', 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM'
    ]
    if not anomalies_tri3.empty:
        export_tri3 = anomalies_tri3[colonnes_export_tri3].rename(columns={'N° Hospit': 'NDA', 'Né(e) le': 'Date naissance', 'Entrée le': 'Date entrée', 'Date Sortie Consolidée': 'Date sortie'})
        
        # Tri par date d'entrée décroissante
        export_tri3['_tri_date'] = pd.to_datetime(export_tri3['Date entrée'], dayfirst=True, errors='coerce')
        export_tri3 = export_tri3.sort_values('_tri_date', ascending=False).drop(columns=['_tri_date'])
    else:
        export_tri3 = pd.DataFrame(columns=['NDA', 'Nom', 'Prénom', 'Date naissance', 'Date entrée', 'Date sortie', 'Date trouvée ailleurs ?', 'Source de la date', 'Exclu', 'Comm. codif. in', 'Comm. ctrl. DIM'])

    # =========================================================
    # EXPORT VERS EXCEL
    # =========================================================
    date_str = datetime.now().strftime("%Y%m%d")
    
    chemin_export_tri1 = dossier_export / f'Tri_1_Ecarts_Hospit_{date_str}.xlsx'
    export_tri1.to_excel(chemin_export_tri1, index=False, engine='openpyxl')
    mettre_en_forme_excel(chemin_export_tri1)
    
    chemin_export_tri2 = dossier_export / f'Tri_2_Ecarts_Seances_{date_str}.xlsx'
    export_tri2.to_excel(chemin_export_tri2, index=False, engine='openpyxl')
    mettre_en_forme_excel(chemin_export_tri2)
    
    chemin_export_tri3 = dossier_export / f'Tri_3_Sans_Date_Sortie_{date_str}.xlsx'
    export_tri3.to_excel(chemin_export_tri3, index=False, engine='openpyxl')
    mettre_en_forme_excel(chemin_export_tri3)
    
    # =========================================================
    # FICHIER DE SYNTHÈSE (4ème fichier)
    # =========================================================
    chemin_synthese = dossier_export / f'Synthese_Conciliation_{date_str}.xlsx'
    
    # Compteurs pour la synthèse
    nb_orbis_total = len(df_orbis_brut)
    nb_orbis_apres_exclusion = len(df_orbis)
    nb_orbis_hospit = len(df_orbis_hospit_sejour)
    nb_orbis_seances = len(df_orbis_seances_filtre)
    nb_hexa_hospit = len(df_hexa_hospit)
    nb_hexa_seances = len(df_hexa_seances)
    
    # Compteurs Tri 1
    nb_tri1_total = len(export_tri1)
    nb_tri1_manque_hexa = len(export_tri1[export_tri1["Origine de l'écart"] == 'Manquant dans Hexagone'])
    nb_tri1_manque_orbis = len(export_tri1[export_tri1["Origine de l'écart"] == 'Manquant dans Orbis'])
    
    # Compteurs Tri 2
    nb_tri2_total = len(export_tri2)
    nb_tri2_manque_hexa = len(export_tri2[export_tri2["Origine de l'écart"] == 'Manquant dans Hexagone'])
    nb_tri2_manque_orbis = len(export_tri2[export_tri2["Origine de l'écart"] == 'Manquant dans Orbis'])
    
    # Compteurs Tri 3
    nb_tri3_total = len(export_tri3)
    nb_tri3_date_trouvee = len(export_tri3[export_tri3['Date trouvée ailleurs ?'] == 'Oui'])
    nb_tri3_date_absente = len(export_tri3[export_tri3['Date trouvée ailleurs ?'] == 'Non'])
    
    synthese_data = {
        'Indicateur': [
            'Date du traitement',
            '---',
            'DONNEES EN ENTREE',
            'Lignes Orbis (brut)',
            'Lignes Orbis (apres exclusion UM 1402)',
            'Lignes Orbis Hospitalisations (sejours uniques)',
            'Lignes Orbis Seances',
            'Lignes Hexagone Hospitalisations',
            'Lignes Hexagone Seances',
            '---',
            'TRI 1 - ECARTS HOSPITALISATIONS',
            'Anomalies totales',
            'Manquant dans Hexagone',
            'Manquant dans Orbis',
            '---',
            'TRI 2 - ECARTS SEANCES',
            'Anomalies totales',
            'Manquant dans Hexagone',
            'Manquant dans Orbis',
            '---',
            'TRI 3 - DATES DE SORTIE MANQUANTES',
            'Anomalies totales',
            'Date trouvee dans un autre fichier',
            'Date absente partout',
        ],
        'Valeur': [
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            '',
            '',
            nb_orbis_total,
            nb_orbis_apres_exclusion,
            nb_orbis_hospit,
            nb_orbis_seances,
            nb_hexa_hospit,
            nb_hexa_seances,
            '',
            '',
            nb_tri1_total,
            nb_tri1_manque_hexa,
            nb_tri1_manque_orbis,
            '',
            '',
            nb_tri2_total,
            nb_tri2_manque_hexa,
            nb_tri2_manque_orbis,
            '',
            '',
            nb_tri3_total,
            nb_tri3_date_trouvee,
            nb_tri3_date_absente,
        ]
    }
    
    df_synthese = pd.DataFrame(synthese_data)
    df_synthese.to_excel(chemin_synthese, index=False, engine='openpyxl')
    mettre_en_forme_excel(chemin_synthese)
    
    logging.info(f"--- TRAITEMENT TERMINÉ ---")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Controle d\'exhaustivite - Conciliation DIM')
    parser.add_argument('--orbis', help='Chemin du fichier Orbis')
    parser.add_argument('--hexa-hospit', help='Chemin du fichier Hexa Hospitalisations')
    parser.add_argument('--hexa-nn', help='Chemin du fichier Hexa Nouveau-Ne')
    parser.add_argument('--hexa-ortho', help='Chemin du fichier Hexa Orthogenie')
    parser.add_argument('--hexa-chimio', help='Chemin du fichier Hexa Chimio')
    parser.add_argument('--hexa-dialyse', help='Chemin du fichier Hexa Dialyse P')
    parser.add_argument('--hexa-hemo', help='Chemin du fichier Hexa Hemodialyse')
    parser.add_argument('--export', help='Dossier d\'export')
    args = parser.parse_args()
    
    # Si des arguments CLI sont fournis, les utiliser
    if args.orbis:
        hexa_hospit = [p for p in [args.hexa_hospit, args.hexa_nn, args.hexa_ortho] if p]
        hexa_seances = [p for p in [args.hexa_chimio, args.hexa_dialyse, args.hexa_hemo] if p]
        lancer_conciliation(
            orbis_path=args.orbis,
            hexa_hospit_paths=hexa_hospit if hexa_hospit else None,
            hexa_seances_paths=hexa_seances if hexa_seances else None,
            export_dir=args.export
        )
    else:
        # Mode par défaut : recherche automatique dans le dossier courant
        lancer_conciliation()